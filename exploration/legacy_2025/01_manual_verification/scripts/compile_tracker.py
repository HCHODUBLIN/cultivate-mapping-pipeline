"""
compile_tracker.py
Compiles manual verification results from city-specific Excel files

Input structure:
- data/bronze/false-positive/*.xlsx (manual review files per city)
  Format: City, Country, Name, URL, Food Sharing Activities, How It Is Shared,
          Date Checked, Comments, Lat, Lon, review
  - Comments empty = Valid FSI
  - Comments filled = False Positive with reason

Output:
- reports/2025_01_manual_verification/manual_verification_results.xlsx
"""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from typing import Dict, List

# Paths
BASE_DIR = Path(__file__).parent.parent.parent.parent.parent
FP_DIR = BASE_DIR / "data" / "bronze" / "false-positive"
OUTPUT_DIR = BASE_DIR / "reports" / "2025_01_manual_verification"
OUTPUT_FILE = OUTPUT_DIR / "manual_verification_results.xlsx"


def categorize_false_positive(comment: str, review: str) -> str:
    """
    Categorize false positive based on Comments and review columns

    Returns category code based on patterns in free text
    """
    if pd.isna(comment) or str(comment).strip() == '':
        return 'VALID'

    # Combine comment and review for pattern matching
    text = f"{str(comment).lower()} {str(review).lower()}"

    # Pattern matching for categories
    if any(x in text for x in ['blog', 'newspaper', 'magazine', 'news', 'media']):
        return 'FP_MEDIA'
    elif any(x in text for x in ['commercial', 'restaurant', 'catering', 'business']):
        return 'FP_COMMERCIAL'
    elif any(x in text for x in ['gov', 'government', 'municipality']):
        return 'FP_GOVERNMENT'
    elif any(x in text for x in ['california', 'ohio', 'wrong location', 'different city']):
        return 'FP_WRONG_LOCATION'
    elif any(x in text for x in ['page not found', 'broken', '404', 'not found']):
        return 'FP_BROKEN_LINK'
    elif any(x in text for x in ['repetition', 'duplicate', 'already listed']):
        return 'FP_DUPLICATE'
    elif any(x in text for x in ['student', 'accommodation', 'university']):
        return 'FP_NON_FSI_ORG'
    elif any(x in text for x in ['supporting', 'support org']):
        return 'FP_SUPPORTING_ORG'
    else:
        return 'FP_OTHER'


def get_fp_category_mapping() -> pd.DataFrame:
    """Get false positive category descriptions"""
    return pd.DataFrame({
        'category': [
            'VALID',
            'FP_MEDIA',
            'FP_COMMERCIAL',
            'FP_GOVERNMENT',
            'FP_WRONG_LOCATION',
            'FP_BROKEN_LINK',
            'FP_DUPLICATE',
            'FP_NON_FSI_ORG',
            'FP_SUPPORTING_ORG',
            'FP_OTHER'
        ],
        'description': [
            'Valid Food Sharing Initiative',
            'Blog, newspaper, magazine, or media coverage',
            'Commercial restaurant or catering business',
            'Government or municipality website',
            'Wrong geographic location (e.g., Dublin CA not Dublin IE)',
            'Page not found or broken link',
            'Duplicate entry (repetition of same FSI)',
            'Non-FSI organization (student housing, etc)',
            'Supporting organization (not actual FSI)',
            'Other false positive reasons'
        ]
    })


def load_city_file(file_path: Path) -> pd.DataFrame:
    """
    Load a single city's manual review Excel file

    Expected columns:
    - City, Country, Name, URL
    - Food Sharing Activities, How It Is Shared
    - Date Checked, Comments, Lat, Lon, review

    Returns:
        DataFrame with: city, name, url, is_valid, fp_category, comments, activities, lat, lon
    """
    try:
        df = pd.read_excel(file_path)

        # Standardize column names (handle variations)
        df.columns = df.columns.str.strip()

        # Determine if valid based on Comments column
        df['is_valid'] = df['Comments'].isna() | (df['Comments'].astype(str).str.strip() == '')

        # Categorize false positives
        df['fp_category'] = df.apply(
            lambda row: categorize_false_positive(row.get('Comments', ''), row.get('review', '')),
            axis=1
        )

        # Extract relevant columns
        result = pd.DataFrame({
            'city': df['City'],
            'name': df['Name'],
            'url': df['URL'],
            'is_valid': df['is_valid'],
            'fp_category': df['fp_category'],
            'comments': df.get('Comments', ''),
            'activities': df.get('Food Sharing Activities', ''),
            'how_shared': df.get('How It Is Shared', ''),
            'date_checked': df.get('Date Checked', ''),
            'lat': df.get('Lat', ''),
            'lon': df.get('Lon', '')
        })

        return result

    except Exception as e:
        print(f"Error loading {file_path.name}: {e}")
        return pd.DataFrame()


def generate_summary_stats(all_data: pd.DataFrame) -> pd.DataFrame:
    """
    Generate summary statistics per city

    Returns:
        DataFrame with: city, total_checked, valid_fsi, false_positives, accuracy
    """
    summary = all_data.groupby('city').agg({
        'url': 'count',  # Total URLs checked
        'is_valid': lambda x: (x == True).sum()  # Valid FSIs
    }).reset_index()

    summary.columns = ['city', 'total_checked', 'valid_fsi']
    summary['false_positives'] = summary['total_checked'] - summary['valid_fsi']
    summary['accuracy_pct'] = (summary['valid_fsi'] / summary['total_checked'] * 100).round(2)

    return summary.sort_values('city')


def generate_fp_analysis(all_data: pd.DataFrame) -> pd.DataFrame:
    """
    Analyze false positive categories

    Returns:
        DataFrame with: category, description, count, percentage
    """
    # Get category mapping
    fp_mapping = get_fp_category_mapping()

    # Filter to false positives only
    fp_data = all_data[all_data['is_valid'] == False].copy()

    # Count by FP category
    fp_counts = fp_data['fp_category'].value_counts().reset_index()
    fp_counts.columns = ['category', 'count']

    # Merge with descriptions
    fp_analysis = fp_mapping.merge(fp_counts, on='category', how='left')
    fp_analysis['count'] = fp_analysis['count'].fillna(0).astype(int)

    # Calculate percentage (excluding VALID)
    total_fp = fp_analysis[fp_analysis['category'] != 'VALID']['count'].sum()
    fp_analysis['percentage'] = (fp_analysis['count'] / total_fp * 100).round(2)
    fp_analysis.loc[fp_analysis['category'] == 'VALID', 'percentage'] = 0

    return fp_analysis.sort_values('count', ascending=False)


def create_tracker_excel(summary: pd.DataFrame, all_data: pd.DataFrame, fp_analysis: pd.DataFrame):
    """
    Create formatted Excel tracker with multiple sheets
    """
    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Sheet 1: Summary statistics by city
        summary.to_excel(writer, sheet_name='Summary', index=False)

        # Sheet 2: Valid FSIs only (most important for research)
        valid_fsis = all_data[all_data['is_valid'] == True].copy()
        valid_fsis[['city', 'name', 'url', 'activities', 'how_shared', 'lat', 'lon']].to_excel(
            writer, sheet_name='Valid FSIs', index=False
        )

        # Sheet 3: False positive analysis
        fp_analysis.to_excel(writer, sheet_name='FP Analysis', index=False)

        # Sheet 4: All validation results (detailed)
        all_data.to_excel(writer, sheet_name='All URLs', index=False)

    # Apply formatting
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(OUTPUT_FILE)


def main():
    """Main compilation workflow"""
    print("=== Manual Verification Tracker Compilation ===\n")

    # Get list of city review files
    if not FP_DIR.exists():
        print(f"Error: False-positive directory not found at {FP_DIR}")
        print("Please ensure data/bronze/false-positive/ exists with city review files")
        return

    city_files = list(FP_DIR.glob("*.xlsx"))

    if not city_files:
        print(f"No Excel files found in {FP_DIR}")
        return

    print(f"Found {len(city_files)} city review files to process\n")

    # Load all city data
    all_data_list = []
    for file_path in city_files:
        print(f"  Loading {file_path.name}...")
        city_data = load_city_file(file_path)
        if not city_data.empty:
            all_data_list.append(city_data)
        else:
            print(f"    ⚠️  Skipped (error or empty)")

    if not all_data_list:
        print("\nError: No data loaded. Check file formats.")
        return

    all_data = pd.concat(all_data_list, ignore_index=True)
    print(f"\n✓ Loaded {len(all_data)} total URLs from {len(all_data_list)} files\n")

    # Generate summary statistics
    print("Generating summary statistics...")
    summary = generate_summary_stats(all_data)
    print(f"  ✓ Summary for {len(summary)} cities")
    print(f"  ✓ Total Valid FSIs: {summary['valid_fsi'].sum()}")
    print(f"  ✓ Total False Positives: {summary['false_positives'].sum()}")

    # Generate FP analysis
    print("\nAnalyzing false positive patterns...")
    fp_analysis = generate_fp_analysis(all_data)
    print(f"  ✓ Categorized into {len(fp_analysis[fp_analysis['count'] > 0])} FP types\n")

    # Create Excel tracker
    print(f"Creating tracker Excel: {OUTPUT_FILE}")
    create_tracker_excel(summary, all_data, fp_analysis)

    print(f"\n✅ Compilation complete!")
    print(f"   Output: {OUTPUT_FILE}")
    print(f"   Total URLs checked: {len(all_data)}")
    print(f"   Valid FSIs: {len(all_data[all_data['is_valid'] == True])} ({(len(all_data[all_data['is_valid'] == True])/len(all_data)*100):.1f}%)")
    print(f"   False Positives: {len(all_data[all_data['is_valid'] == False])} ({(len(all_data[all_data['is_valid'] == False])/len(all_data)*100):.1f}%)")
    print(f"   Cities covered: {all_data['city'].nunique()}")


if __name__ == "__main__":
    main()
